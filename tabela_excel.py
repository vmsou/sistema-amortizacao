import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font


def sheet_header(row_start, column_start):

    n_cell = sheet.cell(row=row_start, column=column_start)
    style_cell(n_cell, 'N', centralized, blue_header_left)

    saldo_cell = sheet.cell(row=row_start, column=column_start + 1)
    style_cell(saldo_cell, 'SALDO DEVEDOR', centralized, title_border)

    amort_cell = sheet.cell(row=row_start, column=column_start + 2)
    style_cell(amort_cell, 'AMORTIZAÇÃO', centralized, title_border)

    juros_cell = sheet.cell(row=row_start, column=column_start + 3)
    style_cell(juros_cell, 'JUROS', centralized, title_border)

    prest_cell = sheet.cell(row=row_start, column=column_start + 4)
    style_cell(prest_cell, 'PRESTAÇÃO', centralized, blue_header_right)


def sheet_total_style(table, row_start, column_start):
    if value := table.limitado:
        wanted_row = row_start + value + table.carencia
    else:
        wanted_row = row_start + table.parcelas + table.carencia + 1

    total_cell_left = sheet.cell(row=wanted_row, column=column_start)
    total_cell_left.border = total_border_left
    total_cell_right = sheet.cell(row=wanted_row, column=column_start + 4)
    total_cell_right.border = total_border_right

    for i in range(column_start + 1, column_start + 4):
        total_cell_base = sheet.cell(row=wanted_row, column=i)
        total_cell_base.border = total_border_base


def adjust_size(worksheet):
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)))) + 0.8
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value


def style_cell(cell, value, align, border, style='Normal'):
    if style == 'Moeda':
        if value != 'x':
            cell.number_format = 'R$ #,##0.00'
    else:
        cell.style = style

    font = Font(name='Calibri', size=11, color='FF000000')
    cell.font = font
    cell.value = value
    cell.alignment = align
    cell.border = border


def run(table, row_start, column_start, filename):
    data = table.create_table()

    header_cell = sheet.cell(row=row_start - 1, column=column_start)
    if (type_name := table.tipo.lower()) == 'sac':
        header_cell.value = type_name.upper()
    else:
        header_cell.value = type_name.title()

    sheet_header(row_start, column_start)

    for key, values in data.items():
        try:
            row = int(key) + column_start + 1
        except:
            row = len(data.keys()) + column_start

        n_cell = sheet.cell(row=row, column=column_start)
        if str(key).isdigit():
            style_cell(n_cell, key, centralized, blue_thick_left)
        else:
            style_cell(n_cell, key.title(), centralized, blue_thick_left)

        saldo_cell = sheet.cell(row=row, column=column_start + 1)
        style_cell(saldo_cell, values['saldo'], centralized, base_border, 'Moeda')

        amort_cell = sheet.cell(row=row, column=column_start + 2)
        style_cell(amort_cell, values['amort'], centralized, base_border, 'Moeda')

        juros_cell = sheet.cell(row=row, column=column_start + 3)
        style_cell(juros_cell, values['juros'], centralized, base_border, 'Moeda')

        prest_cell = sheet.cell(row=row, column=column_start + 4)
        style_cell(prest_cell, values['prest'], centralized, blue_thick_right, 'Moeda')

    sheet_total_style(table, row_start, column_start)
    adjust_size(sheet)

    wb.save(filename=f'{filename}.xlsx')


'''Styles'''
# Alignments
centralized = Alignment(horizontal='center', vertical='center')

# Sides
black_side = Side(style='medium', color='000000')
thin_b_side = Side(style='thin', color='000000')
blue_thick_side = Side(style='thick', color='1F4E78')

# Borders
title_border = Border(top=black_side)
blue_header_left = Border(top=black_side, left=blue_thick_side)
blue_header_right = Border(top=black_side, right=blue_thick_side)
base_border = Border(top=thin_b_side, bottom=thin_b_side)
blue_thick_left = Border(top=thin_b_side, bottom=thin_b_side, left=blue_thick_side)
blue_thick_right = Border(top=thin_b_side, bottom=thin_b_side, right=blue_thick_side)
total_border_base = Border(top=black_side, bottom=thin_b_side)
total_border_left = Border(top=black_side, bottom=thin_b_side, left=blue_thick_side)
total_border_right = Border(top=black_side, bottom=thin_b_side, right=blue_thick_side)


wb = openpyxl.Workbook()
sheet = wb.active
